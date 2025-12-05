import pandas as pd
from pathlib import Path
from admissions import validate_student_id, compute_aggregate, qualify_choice, suggest_programs
from typing import Tuple


def load_data(file_path: Path) -> Tuple[pd.DataFrame,pd.DataFrame,pd.DataFrame]:
    xls = pd.ExcelFile(file_path)
    students = pd.read_excel(xls, sheet_name='students')
    choices = pd.read_excel(xls, sheet_name='choices')
    programs = pd.read_excel(xls, sheet_name='programs')
    return students, choices, programs


def validate_and_clean(students: pd.DataFrame) -> pd.DataFrame:
    # Ensure ID format
    students['ID_valid'] = students['StudentID'].apply(lambda x: validate_student_id(str(x)))
    # Parse subject columns (assume non-ID non-name columns are subjects)
    # If aggregate exists recompute for safety
    subj_cols = [c for c in students.columns if c not in ['StudentID','Name','Aggregate','ID_valid']]
    students['Aggregate'] = students[subj_cols].sum(axis=1)
    return students


def attach_choices_and_evaluate(students: pd.DataFrame, choices: pd.DataFrame, programs: pd.DataFrame) -> pd.DataFrame:
    df = students.merge(choices, on='StudentID', how='left')
    # Build program cutoff lookup
    cutoff_map = dict(zip(programs['Program'], programs['Cutoff']))
    # For each student, determine which choice they get
    def eval_row(row):
        # Determine cutoffs for choices, default to very large negative if program missing
        cut1 = cutoff_map.get(row['Choice1'], -1)
        cut2 = cutoff_map.get(row['Choice2'], -1)
        cut3 = cutoff_map.get(row['Choice3'], -1)
        # Expect cutoffs in descending order for first,second,third
        choice = qualify_choice(row['Aggregate'], [cut1, cut2, cut3])
        return choice
    df['AssignedChoice'] = df.apply(eval_row, axis=1)
    # Create Yes/No columns
    df['FirstChoice'] = df['AssignedChoice'].apply(lambda x: 'Yes' if x == 'First' else 'No')
    df['SecondChoice'] = df['AssignedChoice'].apply(lambda x: 'Yes' if x == 'Second' else 'No')
    df['ThirdChoice'] = df['AssignedChoice'].apply(lambda x: 'Yes' if x == 'Third' else 'No')
    # Suggested program
    programs_list = programs.to_dict('records')
    def suggest(row):
        if row['AssignedChoice'] != 'None':
            return row['Choice1'] if row['FirstChoice']=='Yes' else (row['Choice2'] if row['SecondChoice']=='Yes' else row['Choice3'])
        matches = suggest_programs(row['Aggregate'], programs_list)
        if matches:
            best = matches[-1]
            return best.get('name') or best.get('Program') or best.get('program')
        # No matches: suggest fee-paying option (lowest cutoff program)
        lowest = programs.sort_values('Cutoff', ascending=True).iloc[0]
        return f"Fee-paying option: {lowest['Program']}"
    df['SuggestedProgram'] = df.apply(suggest, axis=1)
    return df


def export_final(df: pd.DataFrame, out_file: Path) -> Path:
    # Export and apply light formatting using openpyxl
    with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='final', index=False)
        summary = df['AssignedChoice'].value_counts().rename_axis('Choice').reset_index(name='Count')
        summary.to_excel(writer, sheet_name='summary', index=False)
    # Apply basic styling
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    wb = load_workbook(out_file)
    ws = wb['final']
    # Header bold
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    # Freeze header row
    ws.freeze_panes = 'A2'
    # Color 'Yes' cells light green for choice columns
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    choice_cols = ['FirstChoice','SecondChoice','ThirdChoice']
    # Build mapping of column name to index
    col_idx = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
    for r in range(2, ws.max_row+1):
        for col in choice_cols:
            if col in col_idx:
                cell = ws.cell(row=r, column=col_idx[col])
                if str(cell.value).strip().lower() == 'yes':
                    cell.fill = green_fill
    wb.save(out_file)
    return out_file

if __name__ == '__main__':
    data_file = Path(__file__).parent / 'admissions_data.xlsx'
    students, choices, programs = load_data(data_file)
    students = validate_and_clean(students)
    df = attach_choices_and_evaluate(students, choices, programs)
    out = Path(__file__).parent / 'admissions_result.xlsx'
    export_final(df, out)
    print('Wrote final result to', out)
